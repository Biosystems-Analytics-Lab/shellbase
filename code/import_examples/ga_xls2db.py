'''
Requirements:
    openpyxl - Unfortunately the xlrd library doesn't handle xlsx files, but openpyxl does.

Command line:
    python ga_xls2db.py --ExcelFile=<path to the datafile>

'''
import sys
import optparse
import traceback
from datetime import datetime
from openpyxl import load_workbook


class Sample:
    def __init__(self):
        self.station_id = None
        self.longitude = None
        self.latitude = None
        self.sample_value = None
        self.tube_code = None
        self.sample_datetime = None
        self.tide_code = None
        self.temperature = None
        self.temp_units = 'celsius'
        self.dissolved_oxygen = None
        self._do_units = 'mg/L'
        self.conductivity = None
        self.conductivity_units = 'mS/cm'
        self.ph = None
        self.salinity = None



def parse_worksheet(xls_filename):
    samples = []
    try:
        #OPen up the excel file as a workbook.
        workbook = load_workbook(filename=xls_filename)
    except Exception as e:
        traceback.print_exc()
    else:
        try:
            #For the time being, we just assume we are interested in the first worksheet.
            worksheet = workbook.active
        except Exception as e:
            traceback.print_exc()
        else:
            for row_ndx, data_row in enumerate(worksheet.iter_rows()):
                #row 0 is the header.
                if row_ndx > 0:
                    print("Processing row: %d" % (row_ndx))
                    try:
                        sample = Sample()
                        sample.station_id = data_row[1].value
                        parts = data_row[0].value.split('/')
                        #Zero pad the date and create a datetime object.
                        date = datetime.strptime('%02d/%02d/%d' % (int(parts[0]), int(parts[1]), int(parts[2])), "%m/%d/%Y")
                        time = data_row[2].value
                        #We combine the individual date and time columns into a python datetime object.
                        sample.sample_datetime = datetime.combine(date, time)
                        sample.latitude = data_row[3].value
                        sample.longitude = data_row[4].value
                        sample.sample_value = data_row[10].value
                        sample.tube_code = data_row[11].value
                        sample.tide_code = data_row[12].value

                        sample.temperature = data_row[5].value
                        sample.dissolved_oxygen = data_row[6].value
                        sample.conductivity = data_row[7].value
                        sample.ph = data_row[8].value
                        sample.salinity = data_row[9].value

                        samples.append(sample)

                    except Exception as e:
                        print("Error on row: %d" % (row_ndx))
                        traceback.print_exc()
    return

def save_to_database(samples):
    return
def main():
    parser = optparse.OptionParser()
    parser.add_option("-c", "--ExcelFile", dest="xl_file", default=None,
                      help="Excel data file to import.")

    (options, args) = parser.parse_args()

    if options.xl_file is None:
        print("No Excel file to process, use --ExcelFile parameter.")
    else:
        print("Excel file: %s to be processed" % (options.xl_file))
        samples = parse_worksheet(options.xl_file)
        save_to_database(samples)
    return

if __name__ == "__main__":
    main()